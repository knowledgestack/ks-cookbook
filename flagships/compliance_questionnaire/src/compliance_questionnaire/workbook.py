"""CAIQ XLSX read/write. Preserves original formatting; fills columns C + E."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# Columns (1-indexed) in the CAIQv4.0.2 sheet.
COL_ID = 1  # Question ID (e.g. A&A-06.2)
COL_QUESTION = 2  # Question text
COL_ANSWER = 3  # CSP CAIQ Answer  ← agent fills
COL_DESCRIPTION = 5  # CSP Implementation Description  ← agent fills
SHEET = "CAIQv4.0.2"
HEADER_ROW = 2  # row 1 is metadata, row 2 is headers; questions start at row 3


@dataclass
class QuestionRow:
    row: int
    control_id: str
    question: str


def load_questions(
    path: Path, *, sheet: str = SHEET, limit: int | None = None
) -> list[QuestionRow]:
    wb = load_workbook(path)
    ws = wb[sheet]
    out: list[QuestionRow] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        control_id = ws.cell(r, COL_ID).value
        question = ws.cell(r, COL_QUESTION).value
        if not control_id or not question:
            continue
        out.append(QuestionRow(row=r, control_id=str(control_id), question=str(question)))
        if limit and len(out) >= limit:
            break
    return out


def write_answers(
    src_path: Path,
    dst_path: Path,
    answers: dict[int, tuple[str, str]],
    *,
    sheet: str = SHEET,
) -> Workbook:
    wb = load_workbook(src_path)
    ws = wb[sheet]
    for row_idx, (answer, description) in answers.items():
        ws.cell(row_idx, COL_ANSWER).value = answer
        ws.cell(row_idx, COL_DESCRIPTION).value = description
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)
    return wb
